import csv
import os
from kivy.uix.image import Image
from kivy.uix.screenmanager import Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.scrollview import ScrollView
from kivy.properties import StringProperty
from kivy.uix.popup import Popup
from kivy_garden.matplotlib import FigureCanvasKivyAgg
import matplotlib.pyplot as plt
import shutil


class ProgramScreen(Screen):
    current_mode = StringProperty("")
    presets = []

    def go_back(self):
        self.manager.transition.direction = 'right'
        self.manager.current = 'second'

    def load_steps_into_screen(self, steps):
        print("Loaded Steps:", steps)
        up_screen = self.manager.get_screen('up_ladder_screen')
        up_screen.set_steps(steps)
        self.manager.transition.direction = 'left'
        self.manager.current = 'up_ladder_screen'

    def open_preset_in_up_ladder(self, preset_steps, graph_source, name, description):
        # Get the 'UpLadderScreen' instance
        up_ladder_screen = self.manager.get_screen('up_ladder_screen')

        # Pass the relevant data to the UpLadderScreen
        up_ladder_screen.graph_source = graph_source  # Path to the graph image
        up_ladder_screen.preset_name = name
        up_ladder_screen.preset_description = description
        up_ladder_screen.load_steps_from_program_screen(preset_steps)  # Pass the steps

        # Transition to the UpLadderScreen
        self.manager.transition.direction = 'left'
        self.manager.current = 'up_ladder_screen'

    def load_presets_from_xlsx(self):
        from openpyxl import load_workbook
        from kivy.uix.boxlayout import BoxLayout
        from kivy.uix.label import Label
        from kivy.uix.progressbar import ProgressBar
        from kivy.uix.popup import Popup
        from kivy.uix.button import Button
        from kivy.uix.scrollview import ScrollView
        from kivy_garden.matplotlib.backend_kivyagg import FigureCanvasKivyAgg
        import matplotlib.pyplot as plt
        import os

        # Show progress popup
        box = BoxLayout(orientation='vertical', padding=20, spacing=10)
        progress_label = Label(text="Loading presets...")
        progress_bar = ProgressBar(max=100, value=0)
        box.add_widget(progress_label)
        box.add_widget(progress_bar)
        loading_popup = Popup(title="Please wait", content=box, size_hint=(0.6, 0.3), auto_dismiss=False)
        loading_popup.open()

        wb = load_workbook("presets/presets.xlsx")
        ws = wb.active

        container = self.ids.presets_container
        container.clear_widgets()

        scrollview = ScrollView(size_hint=(1, 1))
        list_container = BoxLayout(orientation='vertical', size_hint_y=None, spacing=10, padding=10)
        list_container.bind(minimum_height=list_container.setter('height'))

        headers = {cell.value: idx for idx, cell in enumerate(ws[1])}
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        total = len(rows)

        # Clear old graph images
        graph_dir = 'graphs'
        if os.path.exists(graph_dir):
            for file in os.listdir(graph_dir):
                if file.endswith('.png'):
                    os.remove(os.path.join(graph_dir, file))
        else:
            os.makedirs(graph_dir)

        for index, row in enumerate(rows, start=1):
            name = row[headers['name']]
            desc = row[headers['description']]
            temps_raw = row[headers['step_temps']]
            times_raw = row[headers['step_times (min)']]

            if not temps_raw or not times_raw:
                continue

            try:
                temps = [int(t.strip()) for t in str(temps_raw).split(',')]
                times = [int(t.strip()) for t in str(times_raw).split(',')]
            except ValueError:
                print(f"Invalid number format in preset: {name}")
                continue

            if len(temps) != len(times):
                print(f"Mismatch in temps and times count for preset: {name}")
                continue

            steps = [{'temp': temp, 'time': time} for temp, time in zip(temps, times)]

            item_layout = BoxLayout(size_hint_y=None, height=120, spacing=10, padding=5)

            # Generate and save graph
            fig, ax = plt.subplots(figsize=(3, 2))  # Larger figure size for better resolution
            time_points = [0]
            temp_points = []

            for step in steps:
                temp_points.append(step['temp'])
                time_points.append(time_points[-1] + step['time'])

            if temp_points:
                temp_points.append(temp_points[-1])

            ax.step(time_points, temp_points, where='post', linewidth=2)  # Wider line
            ax.set_xticks([])
            ax.set_yticks([])
            ax.axis('off')
            fig.tight_layout()

            graph_path = os.path.join(graph_dir, f'{name}_graph.png')
            fig.savefig(graph_path, dpi=150)  # Higher resolution
            plt.close(fig)

            graph_widget = FigureCanvasKivyAgg(fig)
            graph_widget.size_hint_x = 0.4

            text_button = Button(
                text=f"{name}\n{desc}",
                size_hint_x=0.6,
                halign='left',
                valign='middle'
            )
            text_button.text_size = (text_button.width, None)

            text_button.bind(
                on_release=lambda btn, s=steps, gs=graph_path, n=name, d=desc: self.open_preset_in_up_ladder(s, gs, n,
                                                                                                             d)
            )

            item_layout.add_widget(graph_widget)
            item_layout.add_widget(text_button)
            list_container.add_widget(item_layout)

            # Update progress
            percent = int((index / total) * 100)
            from kivy.clock import Clock
            Clock.schedule_once(lambda dt, p=percent: setattr(progress_bar, 'value', p))

        scrollview.add_widget(list_container)
        container.add_widget(scrollview)

        # Close popup after a tiny delay to ensure it hits 100
        Clock.schedule_once(lambda dt: loading_popup.dismiss(), 0.3)




